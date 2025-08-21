# excel_reporter.py

import sqlite3
import os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment
from typing import List, Dict, Any

import config

logger = config.get_logger(__name__)


# --- ИЗМЕНЕНО: Функция теперь запрашивает данные для ВСЕХ стран сразу ---
def _fetch_and_group_holidays_data(start_date: str, end_date: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    Извлекает данные о праздниках для ВСЕХ стран за указанный период
    и группирует их в словарь по коду страны.
    """
    log_ctx = {'period': f"{start_date} to {end_date}"}
    logger.info("Запрос данных из БД для всех стран за указанный период...", extra={'context': log_ctx})

    query = """
        SELECT
            h.country_code,
            h.holiday_name,
            h.holiday_date,
            r.region_name
        FROM holidays h
        LEFT JOIN regions r ON h.id = r.holiday_id
        WHERE
            h.holiday_date BETWEEN ? AND ?
        ORDER BY
            h.country_code, h.holiday_date, h.holiday_name;
    """
    try:
        with sqlite3.connect(config.DB_PATH) as conn:
            cursor = conn.cursor()
            cursor.execute(query, (start_date, end_date))
            rows = cursor.fetchall()
    except sqlite3.Error as e:
        logger.exception("Ошибка при доступе к БД", extra={'context': log_ctx})
        return {}

    # Агрегируем регионы для каждого уникального праздника (страна, дата, название)
    holidays_aggregated = defaultdict(list)
    for country_code, name, dt, region in rows:
        if region:
            holidays_aggregated[(country_code, name, dt)].append(region)

    # Группируем праздники по странам
    holidays_by_country = defaultdict(list)
    unique_holidays = {(country_code, name, dt) for country_code, name, dt, _ in rows}

    for country_code, name, dt in sorted(list(unique_holidays)):
        regions = holidays_aggregated.get((country_code, name, dt))
        region_str = ", ".join(sorted(regions)) if regions else "Вся страна"

        holidays_by_country[country_code].append({
            "name": name,
            "date": dt,
            "regions": region_str
        })

    logger.info(f"Найдено праздников для {len(holidays_by_country)} стран.", extra={'context': log_ctx})
    # Возвращаем отсортированный по коду страны словарь
    return dict(sorted(holidays_by_country.items()))


def generate_holidays_report(start_date: str, end_date: str) -> str:
    """
    Создает Excel-отчет по праздникам за указанный период на основе всех
    данных в БД и возвращает путь к файлу.
    """
    log_ctx = {'start_date': start_date, 'end_date': end_date, 'report_type': 'excel'}
    logger.info("Начало генерации сводного Excel отчета...", extra={'context': log_ctx})

    if not os.path.exists(config.REPORTS_DIR):
        os.makedirs(config.REPORTS_DIR)

    report_filename = f"holidays_report_{start_date}_to_{end_date}.xlsx"
    report_path = os.path.join(config.REPORTS_DIR, report_filename)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"rep-{start_date}-{end_date}"

    # Стили
    header_font = Font(bold=True, size=12)
    country_font = Font(bold=True, size=11, color="1F497D")  # Сделаем цвет страны другим для наглядности

    # Заголовки и ширина колонок
    headers = ["Страна", "Название праздника", "Дата", "Регионы"]
    sheet.append(headers)
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 50

    # --- ИЗМЕНЕНО: Получаем все данные одним запросом ---
    all_holidays_data = _fetch_and_group_holidays_data(start_date, end_date)
    current_row = 2

    if not all_holidays_data:
        sheet.cell(row=current_row, column=1, value="Праздников за указанный период не найдено ни в одной стране.")
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    else:
        # --- ИЗМЕНЕНО: Итерируемся по данным из БД, а не по config.COUNTRIES ---
        for country_code, holidays in all_holidays_data.items():
            # Записываем название страны и объединяем ячейки для него
            country_cell = sheet.cell(row=current_row, column=1, value=country_code.upper())
            country_cell.font = country_font
            country_cell.alignment = Alignment(vertical='center')
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row + len(holidays) - 1,
                              end_column=1)

            for i, holiday in enumerate(holidays):
                sheet.cell(row=current_row + i, column=2, value=holiday['name'])
                sheet.cell(row=current_row + i, column=3, value=holiday['date'])
                sheet.cell(row=current_row + i, column=4, value=holiday['regions'])

            # Сдвигаем курсор на количество добавленных праздников
            current_row += len(holidays)

    workbook.save(report_path)
    logger.info(f"Excel отчет успешно сохранен: {report_path}", extra={'context': log_ctx})
    return report_path